from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="0F766E")
META_FILL = PatternFill(fill_type="solid", fgColor="F3F4F6")
THIN_SIDE = Side(style="thin", color="D1D5DB")
HEADER_SIDE = Side(style="thin", color="0F766E")
BODY_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER = Border(left=HEADER_SIDE, right=HEADER_SIDE, top=HEADER_SIDE, bottom=HEADER_SIDE)
ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def normalize_excel_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def sanitize_excel_cell(value: object) -> object:
    normalized = normalize_excel_value(value)
    if isinstance(normalized, str):
        return ILLEGAL_XML_CHARS_RE.sub(" ", normalized)
    return normalized


def trim_trailing_empty_cells(row: Iterable[object]) -> list[object]:
    values = [normalize_excel_value(value) for value in row]
    while values and values[-1] == "":
        values.pop()
    return values


def trim_matrix(values: list[list[object]]) -> list[list[object]]:
    rows = [trim_trailing_empty_cells(row) for row in values]
    while rows and not rows[-1]:
        rows.pop()
    return rows


def read_workbook_tables(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False)
    sheets: dict[str, dict[str, object]] = {}
    for worksheet in workbook.worksheets:
        raw_rows: list[list[object]] = []
        for row in worksheet.iter_rows(values_only=True):
            raw_rows.append(list(row))
        values = trim_matrix(raw_rows)
        if not values:
            continue
        headers = [normalize_excel_value(value) for value in values[0]]
        rows: list[list[object]] = []
        for row in values[1:]:
            padded = list(row)
            while len(padded) < len(headers):
                padded.append("")
            rows.append([normalize_excel_value(value) for value in padded[: len(headers)]])
        sheets[worksheet.title] = {"headers": headers, "rows": rows}
    return {"source_xlsx": str(path.resolve()), "sheets": sheets}


def autosize_columns(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        max_length = 0
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            text = "" if value is None else str(value)
            max_length = max(max_length, len(text))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 40)


def style_worksheet(worksheet, headers: list[object]) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Microsoft JhengHei", size=10, color="111827")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BODY_BORDER
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER
        if str(header).startswith("__"):
            for row_index in range(1, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).fill = META_FILL
    autosize_columns(worksheet)


def write_workbook(path: Path, sheet_payloads: list[dict[str, object]]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_payload in sheet_payloads:
        worksheet = workbook.create_sheet(title=str(sheet_payload["name"]))
        headers = [sanitize_excel_cell(value) for value in sheet_payload["headers"]]
        rows = [[sanitize_excel_cell(value) for value in row] for row in sheet_payload["rows"]]
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        style_worksheet(worksheet, headers)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def dump_tables_json(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

