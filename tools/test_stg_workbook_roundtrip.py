from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.export_stg_city_troop_analysis import CITY_FIELD_OFFSETS
from tools.export_stg_workbook import build_workbook_sheets
from tools.import_stg_workbook import apply_city_state_sheet, load_meta, load_record_buffers, rebuild_blob
from tools.stage_ini_excel_codec import read_workbook_tables, write_workbook


class StgWorkbookRoundTripTest(unittest.TestCase):
    """验证 `.stg` 与 Excel 工作簿之间的字节级互转能力。"""

    @classmethod
    def setUpClass(cls) -> None:
        cls.root = Path(__file__).resolve().parents[1]
        cls.stage_path = cls.root / "三国霸业" / "stage01.stg"
        if not cls.stage_path.exists():
            raise unittest.SkipTest("缺少 三国霸业/stage01.stg，跳过 .stg 工作簿互转测试。")

        cls.temp_dir = cls.root / "derived" / "test_tmp"
        cls.temp_dir.mkdir(parents=True, exist_ok=True)
        cls.workbook_path = cls.temp_dir / "stage01_stg_test.xlsx"
        cls.created_paths = [cls.workbook_path, cls.temp_dir / "stage01_stg_population_edit.xlsx"]

        # 导出工作簿较慢，整组测试只执行一次，后续用同一份输入反复验证。
        sheets = build_workbook_sheets(cls.root, "stage01")
        write_workbook(cls.workbook_path, sheets)
        cls.original_blob = cls.stage_path.read_bytes()

    @classmethod
    def tearDownClass(cls) -> None:
        for path in getattr(cls, "created_paths", []):
            if not path.exists():
                continue
            try:
                path.unlink()
            except PermissionError:
                # Windows 偶尔会延迟释放 xlsx 句柄；清理失败不代表互转逻辑失败。
                pass

    def load_workbook_payload(self, path: Path | None = None) -> tuple[dict[str, object], dict[int, bytearray]]:
        """读取测试工作簿，并返回导入器实际使用的 meta 与记录缓冲区。"""

        workbook_payload = read_workbook_tables(path or self.workbook_path)
        sheets = workbook_payload["sheets"]
        meta = load_meta(sheets)
        record_count = int(meta["record_count"])
        buffers = load_record_buffers(sheets, record_count)
        return workbook_payload, buffers

    def test_exported_workbook_has_required_sheets_and_meta(self) -> None:
        workbook_payload = read_workbook_tables(self.workbook_path)
        sheets = workbook_payload["sheets"]

        self.assertEqual(
            {"说明", "meta", "raw_records", "hierarchy_records", "force_city_summary", "city_state", "troop_candidates"},
            set(sheets),
        )
        meta_rows = {str(row[0]): row[1] for row in sheets["meta"]["rows"]}
        self.assertEqual(76, int(meta_rows["stride"]))
        self.assertEqual(2502, int(meta_rows["record_count"]))
        self.assertEqual("010000004c000000", str(meta_rows["header_hex"]))
        self.assertEqual(2502, len(sheets["raw_records"]["rows"]))
        self.assertGreater(len(sheets["city_state"]["rows"]), 0)

    def test_default_import_round_trips_original_bytes(self) -> None:
        workbook_payload, buffers = self.load_workbook_payload()
        patch_report = apply_city_state_sheet(workbook_payload["sheets"], buffers)
        rebuilt = rebuild_blob(load_meta(workbook_payload["sheets"]), buffers)

        self.assertGreater(patch_report["patched_fields"], 0)
        self.assertEqual(self.original_blob, rebuilt)

    def test_raw_only_import_round_trips_original_bytes(self) -> None:
        workbook_payload, buffers = self.load_workbook_payload()
        rebuilt = rebuild_blob(load_meta(workbook_payload["sheets"]), buffers)
        self.assertEqual(self.original_blob, rebuilt)

    def test_city_state_edit_patches_expected_u16_only(self) -> None:
        edited_path = self.temp_dir / "stage01_stg_population_edit.xlsx"
        workbook = load_workbook(self.workbook_path)
        worksheet = workbook["city_state"]
        headers = [cell.value for cell in worksheet[1]]

        population_col = headers.index("candidate_population") + 1
        source_record_col = headers.index("source_record_index") + 1
        city_id_stream_col = headers.index("city_id_stream_index") + 1

        source_record_index = int(worksheet.cell(row=2, column=source_record_col).value)
        city_id_stream_index = int(worksheet.cell(row=2, column=city_id_stream_col).value)
        old_population = int(worksheet.cell(row=2, column=population_col).value)
        new_population = old_population + 1
        worksheet.cell(row=2, column=population_col).value = new_population
        workbook.save(edited_path)
        workbook.close()

        workbook_payload, buffers = self.load_workbook_payload(edited_path)
        apply_city_state_sheet(workbook_payload["sheets"], buffers)
        rebuilt = rebuild_blob(load_meta(workbook_payload["sheets"]), buffers)

        stream_index = city_id_stream_index + CITY_FIELD_OFFSETS["population"]
        absolute_record_index = source_record_index + stream_index // 38
        word_index = stream_index % 38
        file_offset = 8 + absolute_record_index * 76 + word_index * 2

        expected = bytearray(self.original_blob)
        expected[file_offset : file_offset + 2] = new_population.to_bytes(2, "little", signed=False)
        self.assertNotEqual(self.original_blob, rebuilt)
        self.assertEqual(bytes(expected), rebuilt)


if __name__ == "__main__":
    unittest.main()